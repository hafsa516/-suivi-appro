import pandas as pd
import re
import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def trouver_col(df, nom):
    return next((c for c in df.columns if c.strip().upper() == nom.upper()), None)


def week_to_friday(week_str):
    try:
        match = re.search(r'W(\d+)', str(week_str).upper())
        if not match:
            return ''
        week_num = int(match.group(1))
        annee = datetime.datetime.now().year
        vendredi = datetime.datetime.strptime(f'{annee}-W{week_num:02d}-5', '%G-W%V-%u')
        return vendredi.strftime('%d.%m.%Y')
    except:
        return ''


def calculer_confirmation(date_str):
    try:
        if date_str == '':
            return ''
        date_conf = datetime.datetime.strptime(date_str, '%d.%m.%Y')
        aujourd_hui = datetime.datetime.now()
        diff = (date_conf - aujourd_hui).days
        if diff < 0:
            return ''
        elif diff > 30:
            return 'CONFD1'
        elif 15 <= diff <= 30:
            return 'CONFD2'
        else:
            return 'CONFD3'
    except:
        return ''


def traiter_excel(filepath, log_fn, progress_fn):

    log_fn("📂 Lecture du fichier Excel...")
    progress_fn(5)

    try:
        xl = pd.ExcelFile(filepath)
        sheet_names = xl.sheet_names
        log_fn(f"   Feuilles trouvées : {sheet_names}")
    except Exception as e:
        log_fn(f"❌ Impossible de lire le fichier : {e}")
        return

    if 'Commandes' not in sheet_names:
        log_fn("❌ Feuille 'Commandes' introuvable.")
        return
    log_fn("📋 Lecture feuille 'Commandes'...")
    df = pd.read_excel(filepath, sheet_name='Commandes', dtype=str)
    df = df.fillna('').astype(str)
    df.columns = df.columns.str.strip()
    progress_fn(20)

    result = df[
        (df['Date confirmée'].str.strip() == '') &
        (df['Date_Reception'].str.strip() == '') &
        (df['Infos-ach'].str.strip().str.contains(
            r'(?i)cr[eé][eé]\s*par\s*:\s*F[A-Za-z0-9]+', regex=True, na=False))
    ]
    log_fn(f"   Lignes filtrées (Commandes) : {len(result)}")
    progress_fn(35)

    if 'Suivi Appro' not in sheet_names:
        log_fn("❌ Feuille 'Suivi Appro' introuvable.")
        return
    log_fn("📋 Lecture feuille 'Suivi Appro'...")
    df_suivi = pd.read_excel(filepath, sheet_name='Suivi Appro', dtype=str)
    df_suivi = df_suivi.fillna('').astype(str)
    df_suivi.columns = df_suivi.columns.str.strip()
    progress_fn(50)

    result = result.copy()
    result.columns = result.columns.str.strip()
    df_suivi.columns = df_suivi.columns.str.strip()

    result['Liste Cdes'] = result['Liste Cdes'].str.strip().str.lower()
    result['Liste Poste Cdes'] = result['Liste Poste Cdes'].str.strip().str.lower()
    df_suivi['NO DE COMMANDE'] = df_suivi['NO DE COMMANDE'].str.strip().str.lower()
    df_suivi['POSTE CDE'] = df_suivi['POSTE CDE'].str.strip().str.lower()

    cols_result = result.columns.tolist()
    cols_suivi = df_suivi.columns.tolist()

    log_fn("🔗 Fusion des deux feuilles...")
    df_final = result.merge(
        df_suivi,
        left_on=['Liste Cdes', 'Liste Poste Cdes'],
        right_on=['NO DE COMMANDE', 'POSTE CDE'],
        how='left',
        suffixes=('_cmd', '_suivi')
    )
    df_final = df_final.drop(columns=['NO DE COMMANDE', 'POSTE CDE'], errors='ignore')

    cols_a_supprimer = []
    for col_cmd in cols_result:
        for col_s in cols_suivi:
            if col_s in cols_a_supprimer:
                continue
            if col_cmd.strip().lower() != col_s.strip().lower():
                continue
            if col_s not in df_final.columns or col_cmd not in df_final.columns:
                continue
            cols_a_supprimer.append(col_s)
    df_final = df_final.drop(columns=cols_a_supprimer, errors='ignore')
    df_final = df_final.fillna('')
    log_fn(f"   Colonnes finales : {len(df_final.columns)}")
    progress_fn(65)

    col_leadtime = trouver_col(df_final, 'LEADTIME')

    if col_leadtime:
        df_confirmation = df_final[
            df_final[col_leadtime].str.upper().str.contains(r'W\d+', regex=True, na=False)
        ].copy()
        log_fn(f"   Lignes rouges (W+numéro) dans LEADTIME : {len(df_confirmation)}")

        col_cde_cmd   = trouver_col(df_confirmation, 'Doc_achat')
        col_poste_cmd = trouver_col(df_confirmation, 'Poste')
        col_fourn     = trouver_col(df_confirmation, 'Fourn/Div_fourn')
        col_art       = trouver_col(df_confirmation, 'Article')
        col_desig     = trouver_col(df_confirmation, 'Designation')
        col_uac       = trouver_col(df_confirmation, 'UAc')
        col_qte       = trouver_col(df_confirmation, 'A_livrer')

        noms_charge = ['Chargé appro', 'Charge appro', "Chargé d'appro",
                       'Chargé Appro', 'CHARGE APPRO', 'CA', 'CHARGE APPRO_cmd']
        col_charge = next(
            (trouver_col(df_confirmation, n) for n in noms_charge
             if trouver_col(df_confirmation, n) is not None), None)
        if col_charge is None:
            for n in noms_charge:
                candidate = n + '_cmd'
                if candidate in df_confirmation.columns:
                    col_charge = candidate
                    break

        df_out = pd.DataFrame()
        df_out['N°commande'] = df_confirmation[col_cde_cmd].str.strip().str.upper().apply(
            lambda x: x + ',' if x != '' else '') if col_cde_cmd else ''
        df_out['n°poste'] = df_confirmation[col_poste_cmd].str.strip().apply(
            lambda x: x + ',' if x != '' else '') if col_poste_cmd else ''
        df_out['Fournisseur']          = df_confirmation[col_fourn].str.strip()  if col_fourn  else ''
        df_out['référence']            = df_confirmation[col_art].str.strip()    if col_art    else ''
        df_out['Désignation']          = df_confirmation[col_desig].str.strip()  if col_desig  else ''
        df_out['date de confirmation'] = df_confirmation[col_leadtime].apply(week_to_friday)
        df_out['Qte confirmée']        = df_confirmation[col_qte].str.strip()    if col_qte    else ''
        df_out['référence confirmation'] = df_out['date de confirmation'].apply(calculer_confirmation)
        df_out['Unité']                = df_confirmation[col_uac].str.strip()    if col_uac    else ''
        df_out['CA']                   = df_confirmation[col_charge].str.strip() if col_charge else ''

        nb_avant = len(df_out)
        df_out = df_out[df_out['référence confirmation'] != ''].copy()
        log_fn(f"   Lignes supprimées (date vide/passée) : {nb_avant - len(df_out)}")
        df_out = df_out.fillna('').reset_index(drop=True)
    else:
        df_out = pd.DataFrame()
        log_fn("⚠️  Colonne LEADTIME introuvable — tableau confirmation vide")

    progress_fn(80)

    log_fn("💾 Écriture dans le fichier Excel d'origine...")
    wb = load_workbook(filepath)

    for nom in ["Résultat", "Tableau confirmation"]:
        if nom in wb.sheetnames:
            del wb[nom]

    ws_res = wb.create_sheet("Résultat")
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws_res.append(r)

    col_lt_final = trouver_col(df_final, 'LEADTIME')
    if col_lt_final:
        nb_cols = len(df_final.columns)
        for i, val in enumerate(df_final[col_lt_final].astype(str).str.upper(), start=2):
            if val == 'W/O FRNS':
                fill = PatternFill("solid", fgColor="ADD8E6")
            elif re.search(r'W\d+', val):
                fill = PatternFill("solid", fgColor="F29999")
            else:
                continue
            for col_i in range(1, nb_cols + 1):
                ws_res.cell(row=i, column=col_i).fill = fill

    log_fn("✅ Feuille 'Résultat' ajoutée")

    if not df_out.empty:
        ws_conf = wb.create_sheet("Tableau confirmation")
        for r in dataframe_to_rows(df_out, index=False, header=True):
            ws_conf.append(r)
        log_fn(f"✅ Feuille 'Tableau confirmation' ajoutée ({len(df_out)} lignes)")
    else:
        log_fn("⚠️  Tableau confirmation vide — feuille non créée")

    wb.save(filepath)
    log_fn(f"💾 Fichier sauvegardé : {os.path.basename(filepath)}")
    progress_fn(100)
    log_fn("🎉 Traitement terminé avec succès !")